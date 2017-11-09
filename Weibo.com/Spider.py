import Agents
import random
import requests
import time
from bs4 import BeautifulSoup
import os
import re
import json
import openpyxl
import Ban


class WbSpider(object):
    def __init__(self, cookie=''):
        self.__user_agent = Agents.agents
        self.__cookies = cookie
        self.__headers = {
            "User_Agent": random.choice(self.__user_agent),
            "cookie": random.choice(self.__cookies)
        }
        self.__url =( "http://s.weibo.com/weibo/%25E7%2599%25BB%25E5%25B0%258"
                      "1&region=custom:41:1000&scope=ori&suball=1&page={page}")
        self.__html = []

        self.__path_page_raw = r"./data/text/page_raw"
        self.__path_html = r"./data/text/html"
        self.__path_work_sheet = r"./data/Result/post.xlsx"

    def set_cookie(self, cookie):
        self.__cookies = cookie
        self.__headers["cookie"] = self.__cookies

    def set_url(self, url):
        self.__url = url

    def get_url(self):
        return self.__url

    def get_page_raw(self):
        # for i in range(1, 51):
        for i in range(14, 15):
            r = requests.get(self.__url.format(page=i), headers=self.__headers)
            r.encoding = "UTF-8"
            with open(self.__path_page_raw + r"/page" + str(i) + ".txt", 'w', encoding="UTF-8") as f:
                f.write(r.text)
            print("get_raw_page():_____获得第 " + str(i) + "/50 页源码。_____")
            self.__headers["cookie"] = random.choice(self.__cookies)
            time.sleep(random.randrange(3, 7, 2))

    def page_filter(self):
        list_raw = os.listdir(self.__path_page_raw)
        pattern = re.compile(r'\(({"pid":"pl_weibo_direct".*)\)</script>')

        index = 0
        for raw in list_raw:
            index += 1
            with open(self.__path_page_raw + r'/' + raw, 'r', encoding="UTF-8") as f1:
                with open(self.__path_html + r'/html_' + str(index) + r".txt", 'w', encoding="UTF-8") as f2:
                    mo = pattern.search(f1.read())
                    if mo is not None:
                        j_text = mo.group(1)
                        html = json.loads(j_text)['html']
                        f2.write(html)

    def load_html(self):
        list_html = os.listdir(self.__path_html)
        for html in list_html:
            with open(self.__path_html + r'/' + html, 'r', encoding="UTF-8") as f:
                self.__html.append(f.read())
        print("load_html():________载入HTML代码完成!________")

    def extract_items_from_html(self):
        if len(self.__html) is 0:
            self.load_html()

        # 过滤规则
        pattern_ban = re.compile(Ban.ban_dengfeng, re.VERBOSE | re.IGNORECASE | re.DOTALL)

        # 如果没有工作表，则创建
        if not os.path.exists(self.__path_work_sheet):
            wb_tmp = openpyxl.Workbook()
            wb_tmp.create_sheet()
            wb_tmp.save(self.__path_work_sheet)

        # 将微博数据存储到excel工作簿中
        wb = openpyxl.load_workbook(self.__path_work_sheet)
        sheet = wb.get_active_sheet()
        sheet["A1"] = "User ID"
        sheet["B1"] = "User Post"
        sheet["C1"] = "Time"
        sheet["D1"] = "Num-like"
        sheet["E1"] = "Num-forward"
        sheet["F1"] = "Num-comment"
        index = 1

        # 外层循环，页面。
        for html in self.__html:
            soup = BeautifulSoup(html, "lxml")
            # 用户名 content.attrs["nick-name"]
            # 博文内容
            tag_content = soup.select('p[class="comment_txt"]')
            # 时间，格式化的time获取 : time.attrs["title"]
            tag_time = soup.select('a[node-type="feed_list_item_date"]')
            # 转发数
            tag_num_forward = soup.select('a[action-type="feed_list_forward"] span em')
            # 评论数 正则表达式获取！！！！（将评论两个字过滤）
            tag_num_comment = soup.select('a[action-type="feed_list_comment"] span')
            # 赞数
            tag_num_like = soup.select('a[action-type="feed_list_like"] span em')

            # 里层循环，微博条目。
            for j in range(len(tag_content)):
                webo_content = tag_content[j].getText()

                # 如果博文内容中出现屏蔽词
                if pattern_ban.search(webo_content) is not None:
                    continue

                nick_name = tag_content[j].attrs["nick-name"]
                webo_time = tag_time[j].attrs["title"]
                num_forward = tag_num_forward[j].getText()
                num_comment = tag_num_comment[j].getText()[2:]
                num_like = tag_num_like[j].getText()

                sheet["A" + str(index + 1)] = nick_name
                sheet["B" + str(index + 1)] = webo_content
                sheet["C" + str(index + 1)] = webo_time

                if num_forward is '':
                    sheet["D" + str(index + 1)] = 0
                else:
                    sheet["D" + str(index + 1)] = int(num_forward)

                if num_comment is '':
                    sheet["E" + str(index + 1)] = 0
                else:
                    sheet["E" + str(index + 1)] = int(num_comment)

                if num_like is '':
                    sheet["F" + str(index + 1)] = 0
                else:
                    sheet["F" + str(index + 1)] = int(num_like)

                index += 1

        wb.save(self.__path_work_sheet)
        print("extract_items_from_html():________" + str(index) + "条微博数据已保存至工作簿！________")


