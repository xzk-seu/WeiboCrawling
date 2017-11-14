from Avenger import Avengers
import json
import re
from bs4 import BeautifulSoup
import os
import openpyxl
import dict.Ban


# 收集登封——此地热议
class Captain(Avengers):
    def __init__(self):
        super(Captain, self).__init__()
        self.path_work_sheet = r'./data/Result/Post_dengfeng.xlsx'

    # 将主页包含的HTML记录下来
    def attack_home_page(self):
        pattern = re.compile(
            "FM\.view\(({\"ns\":\"pl\.content\.homeFeed\.index\",\"domid\":\"Pl_Core_MixedFeed__35\",.*)\)</script>")
        self.get_html_index()
        for index in range(1, self.max_page_num['dengfeng'] + 1):
            r = self.get_page(self.urls['dengfeng'].format(page=index))
            mo = pattern.search(r)
            if mo is not None:
                js_text = mo.group(1)
            else:
                print('attack_home_page():_____fail to attack page', str(index), '!_____')
                continue
            page = json.loads(js_text)['html']
            with open(self.path_html + '/html_' + str(self.html_index) + '.txt', 'w', encoding='utf-8') as f:
                f.write(page)
                self.html_index += 1

        print('attack_home_page():________Done!________')

    # 将动态页面的html记录下来
    def attack_page_bar(self):
        self.get_html_index()
        url = self.urls['dengfengJS']
        for i in range(1, self.max_page_num['dengfeng'] + 1):
            for j in range(0, 2):
                r = self.get_page(url.format(page=i, pre_page=i, pagebar=j, current_page=(i*3-2+j)))
                page = json.loads(r)['data']
                with open(self.path_html + '/html_' + str(self.html_index) + '.txt', 'w', encoding='utf-8') as f:
                    f.write(page)
                    self.html_index += 1
        print("attack_page_bar():________Done!________")

    def __load_html(self):
        list_html = os.listdir(self.path_html)
        for html in list_html:
            with open(self.path_html + r'/' + html, 'r', encoding="UTF-8") as f:
                self.html.append(f.read())
        print("load_html():________载入HTML代码完成!________")

    def extract_items_from_html(self):
        if len(self.html) is 0:
            self.__load_html()

        # 过滤规则
        pattern_ban = re.compile(dict.Ban.ban_dengfeng, re.VERBOSE | re.IGNORECASE | re.DOTALL)

        # 如果没有工作表，则创建
        if not os.path.exists(self.path_work_sheet):
            wb_tmp = openpyxl.Workbook()
            wb_tmp.create_sheet()
            wb_tmp.save(self.path_work_sheet)

        # 将微博数据存储到excel工作簿中
        wb = openpyxl.load_workbook(self.path_work_sheet)
        sheet = wb.get_active_sheet()
        sheet["A1"] = "User ID"
        sheet["B1"] = "User Post"
        sheet["C1"] = "Time"
        sheet["D1"] = "Num-like"
        sheet["E1"] = "Num-forward"
        sheet["F1"] = "Num-comment"
        index = 1

        # 外层循环，页面。
        for html in self.html:
            soup = BeautifulSoup(html, "lxml")
            # 用户名
            tag_nick_name = soup.select('a[class="W_f14 W_fb S_txt1"]')
            # 博文内容
            tag_content = soup.select('div[node-type="feed_list_content"]')
            # 时间，格式化的time获取 : time.attrs["title"]
            tag_time = soup.select('a[node-type="feed_list_item_date"]')
            # 转发数
            tag_num_forward = soup.select('span[node-type="forward_btn_text"] span')
            # 评论数 （将评论两个字过滤）
            tag_num_comment = soup.select('span[node-type="comment_btn_text"] span')
            # 赞数
            tag_num_like = soup.select('span[node-type="like_status"]')

            # 里层循环，微博条目。
            for j in range(len(tag_content)):
                webo_content = tag_content[j].getText().strip()

                # 如果博文内容中出现屏蔽词
                if pattern_ban.search(webo_content) is not None:
                    continue

                nick_name = tag_nick_name[j].getText().strip()
                webo_time = tag_time[j].attrs["title"]
                num_forward = tag_num_forward[j].getText().replace('\u8f6c\u53d1', '').replace('\ue607', '')
                num_comment = tag_num_comment[j].getText().replace('\u8bc4\u8bba', '').replace('\ue608', '')
                num_like = tag_num_like[j].getText()[1:].replace('\u8d5e', '')

                sheet["A" + str(index + 1)] = nick_name
                sheet["B" + str(index + 1)] = webo_content
                sheet["C" + str(index + 1)] = webo_time

                if num_forward is '':
                    sheet["D" + str(index + 1)] = 0
                else:
                    sheet["D" + str(index + 1)] = int(num_forward)

                if num_comment is '':
                    sheet["E" + str(index + 1)] = 0
                else:
                    sheet["E" + str(index + 1)] = int(num_comment)

                if num_like is '':
                    sheet["F" + str(index + 1)] = 0
                else:
                    sheet["F" + str(index + 1)] = int(num_like)

                index += 1

        wb.save(self.path_work_sheet)
        print("extract_items_from_html():________" + str(index) + "条微博数据已保存至工作簿！________")


if __name__ == '__main__':
    c = Captain()
    c.attack_home_page()
    c.attack_page_bar()
    c.__load_html()
    c.extract_items_from_html()
