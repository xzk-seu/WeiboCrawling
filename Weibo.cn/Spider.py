import requests
import random
import time
import bs4
import os
import openpyxl
import re
import logging
logging.basicConfig(filename="./data/logs/log.txt", level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program.')


class WeiboSpider(object):
    def __init__(self, cookie):
        self.__cookie = cookie
        self.__user_agents = [
            'Mozilla/5.0 (iPhone; CPU iPhone OS 9_1 like Mac OS X) \
            AppleWebKit/601.1.46 (KHTML, like Gecko) Version/9.0 '
            'Mobile/13B143 Safari/601.1]',
            'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/48.0.2564.23 Mobile Safari/537.36',
            'Mozilla/5.0 (Linux; Android 5.1.1; Nexus 6 Build/LYZ28E) AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/48.0.2564.23 Mobile Safari/537.36'
        ]
        self.__headers ={
            "User_Agent": random.choice(self.__user_agents),
            "cookie": self.__cookie
        }
        # 设定获取获取的最大微博页面数
        self.__MAX_PAGE = 10
        # 默认搜索登封，url后接页码可跳转到相应页面
        self.__url_search = "https://weibo.cn/search/mblog?hideSearchFrame=&keyword=%E7%99%BB%E5%B0%81&page="
        self.__list_raw_page = []

        self.__path_dir_back_up = r"./data/raw_page"
        self.__path_work_sheet = r"./data/Result/post.xlsx"


    def set_cookie(self, cookie):
        self.__cookie = cookie
        self.__headers["cookie"] = cookie

    def set_max_page(self, max_page):
        self.__MAX_PAGE = max_page

    def set_url_search(self, url_s):
        self.__url_search = url_s

    def get_headers(self):
        return self.__headers

    # 下载所有页面源码
    def get_raw_wb(self):
        this_page = 1
        while this_page <= self.__MAX_PAGE:
            url = self.__url_search + str(this_page)
            r = requests.get(url, headers=self.__headers)
            r.encoding = "utf-8"
            self.__list_raw_page.append(r.text)
            with open(r"./data/raw_page/rp" + str(this_page) +".txt", 'w', encoding="utf-8") as f:
                f.write(r.text)
            print("get_raw_wb():_____第 " + str(this_page)+ "/" + str(self.__MAX_PAGE) + " 页源码下载完成_____")
            this_page += 1
            time.sleep(3)
        print("get_raw_wb():________源码全部下载完成________")

    def get_raw_wb_from_backup(self):
        list_backup = os.listdir(self.__path_dir_back_up)
        num_all_file = len(list_backup)
        this_file = 1
        for name in list_backup:
            with open(self.__path_dir_back_up + r'/' + name, 'r', encoding="UTF-8") as f:
                self.__list_raw_page.append(f.read())
            print("get_raw_wb_from_backup():_____第 " + str(this_file) + r'/' + str(num_all_file) + " 页源码载入完成_____")
            this_file += 1
        print("get_raw_wb_from_backup():________源码全部载入完成________")

    def extract_wb(self):
        num_all_page = len(self.__list_raw_page)
        if num_all_page == 0:
            print("wb_content_filter(): 源码列表为空，请重新下载或从备份中读取")
            return

        if not os.path.exists(self.__path_work_sheet):
            open(self.__path_work_sheet, 'w').close()
            wb_tmp = openpyxl.Workbook()
            wb_tmp.create_sheet()
            wb_tmp.save(self.__path_work_sheet)

        # 将微博数据存储到excel工作簿中
        wb = openpyxl.load_workbook(self.__path_work_sheet)
        sheet = wb.get_active_sheet()

        # this_post:微博条数索引
        this_post = 2

        # 计算工作表的起始位置
        # for rowNum in range(2, sheet.get_highest_row()):
        #     user_id = sheet.cell(row=rowNum, column=1)
        #     if user_id == "":
        #         content = sheet.cell(row=rowNum, column=2)
        #         if content == "":
        #             this_post = rowNum
        #             break

        # 提取每个页面中的元素
        pattern_good = re.compile(r"st=d525c7\">\u8D5E\[(\d+)\]")
        pattern_repost = re.compile(r"rl=1\">\u8F6C\u53D1\[(\d+)\]")
        pattern_comment = re.compile(r"class=\"cc\">\u8BC4\u8BBA\[(\d+)\]")

        #屏蔽词
        pattern_ban = re.compile(r'''(magazine)|
                                    (Bazaar)|
                                    (\u6742\u5fd7)|  # 杂志
                                    (\u6708\u520a)|  # 月刊
                                    (\u6613\u70ca\u5343\u73ba)| # 易烊千玺
                                    (\u5199\u771f)| # 写真
                                    (\u738b\u6e90)| # 王源
                                    (\u5434\u4ea6\u51e1)| #吴亦凡
                                    (\u520a\u7269)| # 刊物
                                    (\u82ad\u838e)| # 芭莎
                                    (\u5c01\u9762) # 封面
                                    ''', re.VERBOSE | re.IGNORECASE | re.DOTALL)

#         pattern_ban = re.compile(r'''12312312312312312312''', re.VERBOSE)

        for page in self.__list_raw_page:
            soup = bs4.BeautifulSoup(page, "lxml")
            
            # 用户id
            tag_nk = soup.select('div a[class="nk"]')
            # 用户发表内容
            tag_ctt = soup.select('div span[class="ctt"]')
            # 时间戳
            tag_ct = soup.select('div span[class="ct"]')
            # 赞数
            list_n_good = pattern_good.findall(page)
            # 转发数
            list_n_repost = pattern_repost.findall(page)
            # 评论数
            list_n_comment = pattern_comment.findall(page)

            index = 0
            for item in tag_nk:
                ban = pattern_ban.search(page)
                if ban is not None:
                    logging.debug("ban:" + ban.group())
                    index += 1
                    continue
                else:
                    sheet["A" + str(this_post)] = item.getText()
                    sheet["B" + str(this_post)] = tag_ctt[index].getText()[1:]
                    sheet["C" + str(this_post)] = tag_ct[index].getText()
                    sheet["D" + str(this_post)] = int(list_n_good[index])
                    sheet["E" + str(this_post)] = int(list_n_repost[index])
                    sheet["F" + str(this_post)] = int(list_n_comment[index])
                    this_post += 1
                    index += 1

        wb.save(self.__path_work_sheet)
        print("extract_wb():________博文已写入到工作簿________")

if __name__ == "__main__":
    cookie = ""
    wbSpider = WeiboSpider(cookie=cookie)
    # wbSpider.set_max_page(90)
    # wbSpider.get_raw_wb()
    wbSpider.get_raw_wb_from_backup()
    wbSpider.extract_wb()

    logging.debug('End of program.')
