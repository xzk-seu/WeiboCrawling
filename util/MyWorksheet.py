import os
import openpyxl


class MyWorksheet:
    def __init__(self):
        self.__path_worksheet = os.path.join(os.getcwd(), 'RESULTS', 'sheet.xlsx')

    def save_worksheet(self, posts):
        # 如果没有工作表，则创建
        if not os.path.exists(self.__path_worksheet):
            wb_tmp = openpyxl.Workbook()
            wb_tmp.create_sheet()
            wb_tmp.save(self.__path_worksheet)

            # 将微博数据存储到excel工作簿中
            wb = openpyxl.load_workbook(self.__path_worksheet)
            sheet = wb.get_active_sheet()
            sheet["A1"] = "User ID"
            sheet["B1"] = "User Post"
            sheet["C1"] = "Time"
            sheet["D1"] = "Num-like"
            sheet["E1"] = "Num-forward"
            sheet["F1"] = "Num-comment"

            index = 1
            for post in posts:
                sheet["A" + str(index + 1)] = post[0]  # nick_name
                sheet["B" + str(index + 1)] = post[1]  # content
                sheet["C" + str(index + 1)] = post[2]  # time
                sheet["D" + str(index + 1)] = post[3]  # num_forward
                sheet["E" + str(index + 1)] = post[4]  # num_comment
                sheet["F" + str(index + 1)] = post[5]  # num_like
                index += 1

            wb.save(self.__path_worksheet)
            print("extract_items_from_html():________" + str(index) + "条微博数据已保存至工作簿！________")
